from __future__ import annotations

import json
import re
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = ROOT.parent / "Velogames Saison 2026 - Classements.xlsx"
LOGOS_DIR = ROOT.parent / "Logos-Courses"
OUTPUT_JSON = ROOT / "assets" / "data" / "velogames-data.json"
OUTPUT_JS = ROOT / "assets" / "data" / "velogames-data.js"
SITE_LOGOS_DIR = ROOT / "assets" / "course-logos"

COURSE_DISPLAY_NAMES = {
    "Down Under": "Santos Tour Down Under",
    "Itzulia": "Itzulia Basque Country",
    "Romandie": "Tour de Romandie",
    "Vuelta Fem": "Vuelta Femenina",
    "Critérium": "Critérium du Dauphiné",
    "Tour Femmes": "Tour de France Femmes avec Zwift",
    "Tour Romandie Femmes": "Tour de Romandie Féminin",
}


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_value = ascii_value.lower().replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", "-", ascii_value).strip("-")


def clean_number(value):
    if value is None:
        return None
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return round(value, 2)
    return value


def get_display_course_name(value: str) -> str:
    return COURSE_DISPLAY_NAMES.get(value, value)


def build_logo_map() -> dict[str, str]:
    mapping = {}
    if not LOGOS_DIR.exists():
        return mapping

    SITE_LOGOS_DIR.mkdir(parents=True, exist_ok=True)

    for path in LOGOS_DIR.iterdir():
        if not path.is_file():
            continue
        destination = SITE_LOGOS_DIR / path.name
        if path.resolve() != destination.resolve():
            shutil.copy2(path, destination)
        slug = slugify(path.stem)
        mapping[slug] = f"./assets/course-logos/{path.name}"

    aliases = {
        "down-under": "tour-down-under",
        "santos-tour-down-under": "tour-down-under",
        "tirreno-adriatico": "tirreno-adriatico",
        "tirreno-adriatico": "tirrenoadriatico",
        "itzulia": "itzulia-basque-country",
        "romandie": "tour-de-romandie",
        "tour-romandie-femmes": "tour-de-romandie-femmes",
        "tour-de-romandie-feminin": "tour-de-romandie-femmes",
        "romandie-femmes": "tour-de-romandie-femmes",
        "vuelta-fem": "lvf23-logo-positivo-color-rgb",
        "vuelta-femenina": "lvf23-logo-positivo-color-rgb",
        "la-vuelta-femenina": "lvf23-logo-positivo-color-rgb",
        "tour-femmes": "tour-de-france-femmes-avec-zwift",
        "criterium": "criterium-du-dauphine",
        "criterium-du-dauphine": "criterium-du-dauphine",
        "simac-ladies-tour": "simac-ladies-tour",
    }
    for alias, target in aliases.items():
        if target in mapping:
            mapping[alias] = mapping[target]

    return mapping


def read_global_ranking(details_sheet):
    ranking = []
    for row in range(24, 40):
        rank = details_sheet[f"F{row}"].value
        name = details_sheet[f"G{row}"].value
        points = details_sheet[f"H{row}"].value
        if rank and name:
            ranking.append(
                {
                    "rank": int(rank),
                    "name": str(name),
                    "points": clean_number(points) or 0,
                }
            )
    return ranking


def read_team_ranking(details_sheet):
    teams = []
    for row in range(24, 27):
        rank = details_sheet[f"J{row}"].value
        name = details_sheet[f"K{row}"].value
        points = details_sheet[f"L{row}"].value
        if rank and name:
            teams.append(
                {
                    "rank": int(rank),
                    "name": str(name),
                    "points": clean_number(points) or 0,
                }
            )
    return teams


def read_course_headers(details_sheet):
    courses = []
    for col in range(3, 41, 2):
        name = details_sheet.cell(2, col).value
        if name:
            courses.append({"name": str(name), "slug": slugify(str(name))})
    return courses


def read_progression(details_sheet, logo_map):
    players = []
    courses = []
    detail_course_cols = list(range(3, 41, 2))
    for idx, col in enumerate(range(3, 22)):
        course_name = details_sheet.cell(43, col).value
        if course_name:
            course_slug = slugify(str(course_name))
            detail_rank_col = detail_course_cols[idx]
            detail_points_col = detail_rank_col + 1
            is_played = False
            for row in range(4, 20):
                rank_value = details_sheet.cell(row, detail_rank_col).value
                points_value = details_sheet.cell(row, detail_points_col).value
                if rank_value not in (None, "") or clean_number(points_value) not in (None, 0):
                    is_played = True
                    break
            courses.append(
                {
                    "name": str(course_name),
                    "slug": course_slug,
                    "logo": logo_map.get(course_slug),
                    "played": is_played,
                }
            )

    for row in range(44, 60):
        name = details_sheet.cell(row, 2).value
        if not name:
            continue
        totals = []
        for col in range(3, 22):
            totals.append(clean_number(details_sheet.cell(row, col).value) or 0)
        players.append({"name": str(name), "totals": totals})

    rankings_by_course = []
    for idx, course in enumerate(courses):
        standing = []
        for player in players:
            standing.append({"name": player["name"], "points": player["totals"][idx]})
        standing.sort(key=lambda item: item["points"], reverse=True)
        ranking_map = {entry["name"]: position + 1 for position, entry in enumerate(standing)}
        rankings_by_course.append(
            {
                "course": course["name"],
                "slug": course["slug"],
                "ranks": ranking_map,
            }
        )

    return {"courses": courses, "players": players, "rankingsByCourse": rankings_by_course}


def read_palmares_from_details(details_sheet, logo_map):
    entries = []
    course_blocks = []
    for col in range(3, 41, 2):
        course_name = details_sheet.cell(2, col).value
        if not course_name:
            continue
        course_blocks.append(
            {
                "name": str(course_name),
                "rankCol": col,
                "pointsCol": col + 1,
            }
        )

    for block in course_blocks:
        course = block["name"]
        official_name = get_display_course_name(course)
        slug = slugify(official_name)
        numeric_rankings = []
        has_data = False

        for row in range(4, 20):
            player = details_sheet.cell(row, 2).value
            if not player:
                continue
            raw_rank = details_sheet.cell(row, block["rankCol"]).value
            points = clean_number(details_sheet.cell(row, block["pointsCol"]).value) or 0

            if raw_rank not in (None, "") or points != 0:
                has_data = True

            if isinstance(raw_rank, (int, float)):
                numeric_rankings.append((int(raw_rank), str(player)))

        numeric_rankings.sort(key=lambda item: item[0])

        entries.append(
            {
                "course": official_name,
                "slug": slug,
                "logo": logo_map.get(slug) or logo_map.get(slugify(course)),
                "first": numeric_rankings[0][1] if len(numeric_rankings) >= 1 else None,
                "second": numeric_rankings[1][1] if len(numeric_rankings) >= 2 else None,
                "third": numeric_rankings[2][1] if len(numeric_rankings) >= 3 else None,
                "last": numeric_rankings[-1][1] if numeric_rankings else None,
                "isPlayed": has_data,
            }
        )
    return entries


def read_points_rules(points_sheet):
    categories = []
    for col in range(4, 10):
        category = points_sheet.cell(3, col).value
        if not category:
            continue
        ranks = []
        for row in range(4, 24):
            rank = points_sheet.cell(row, 3).value
            points = points_sheet.cell(row, col).value
            if rank is None or points is None:
                continue
            ranks.append({"rank": int(rank), "points": clean_number(points)})
        categories.append({"name": str(category), "ranks": ranks})
    return categories


def read_details(details_sheet):
    player_rows = []
    course_blocks = []
    for col in range(3, 41, 2):
        course_name = details_sheet.cell(2, col).value
        if course_name:
            course_blocks.append(
                {
                    "name": str(course_name),
                    "rankCol": col,
                    "pointsCol": col + 1,
                }
            )

    for row in range(4, 20):
        player = details_sheet.cell(row, 2).value
        if not player:
            continue
        results = []
        for block in course_blocks:
            results.append(
                {
                    "course": block["name"],
                    "rank": details_sheet.cell(row, block["rankCol"]).value,
                    "points": clean_number(details_sheet.cell(row, block["pointsCol"]).value)
                    or 0,
                }
            )
        total = clean_number(details_sheet.cell(row, 42).value) or 0
        player_rows.append({"name": str(player), "total": total, "results": results})
    return player_rows


def build_payload():
    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    details_sheet = workbook[workbook.sheetnames[2]]
    points_sheet = workbook[workbook.sheetnames[3]]
    logo_map = build_logo_map()

    source_mtime = datetime.fromtimestamp(SOURCE_XLSX.stat().st_mtime)
    courses = read_course_headers(details_sheet)
    progression = read_progression(details_sheet, logo_map)
    played_courses_count = sum(
        1 for course in progression["courses"] if course["played"]
    )

    return {
        "meta": {
            "siteTitle": "Classement VELOGAMES - Saison 2026",
            "sourceFile": SOURCE_XLSX.name,
            "lastUpdated": source_mtime.isoformat(timespec="seconds"),
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "playersCount": len(read_global_ranking(details_sheet)),
            "coursesCount": len(courses),
            "playedCoursesCount": played_courses_count,
        },
        "home": {
            "globalRanking": read_global_ranking(details_sheet),
            "teamRanking": read_team_ranking(details_sheet),
            "progression": progression,
        },
        "palmares": read_palmares_from_details(details_sheet, logo_map),
        "details": read_details(details_sheet),
        "pointsRules": read_points_rules(points_sheet),
    }


def main():
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    payload = build_payload()
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_JS.write_text(
        "window.VELOGAMES_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    print(f"JSON exported to {OUTPUT_JSON}")
    print(f"JS exported to {OUTPUT_JS}")


if __name__ == "__main__":
    main()
